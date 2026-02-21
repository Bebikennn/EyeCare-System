"""
Export utilities for data export functionality
"""
import csv
import io
from datetime import datetime
from flask import Response
import json


def export_to_csv(data, columns, filename=None):
    """
    Export data to CSV format
    
    Args:
        data: List of dictionaries or objects with to_dict method
        columns: List of column names or dict {header: key}
        filename: Optional filename for download
    
    Returns:
        Flask Response with CSV data
    """
    # Create CSV in memory
    output = io.StringIO()
    
    # Handle column format
    if isinstance(columns, dict):
        headers = list(columns.keys())
        keys = list(columns.values())
    else:
        headers = columns
        keys = columns
    
    writer = csv.writer(output)
    writer.writerow(headers)
    
    # Write data rows
    for item in data:
        if hasattr(item, 'to_dict'):
            item = item.to_dict()
        
        row = []
        for key in keys:
            value = item.get(key, '')
            # Handle datetime objects
            if isinstance(value, datetime):
                value = value.isoformat()
            # Handle None
            if value is None:
                value = ''
            row.append(str(value))
        
        writer.writerow(row)
    
    # Create response
    output.seek(0)
    response = Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename={filename or "export.csv"}'
        }
    )
    
    return response


def export_to_json(data, filename=None):
    """
    Export data to JSON format
    
    Args:
        data: List of dictionaries or objects with to_dict method
        filename: Optional filename for download
    
    Returns:
        Flask Response with JSON data
    """
    # Convert to dict if needed
    json_data = []
    for item in data:
        if hasattr(item, 'to_dict'):
            json_data.append(item.to_dict())
        else:
            json_data.append(item)
    
    # Create response
    response = Response(
        json.dumps(json_data, indent=2, default=str),
        mimetype='application/json',
        headers={
            'Content-Disposition': f'attachment; filename={filename or "export.json"}'
        }
    )
    
    return response


def export_to_excel(data, columns, filename=None):
    """
    Export data to Excel format (requires openpyxl)
    
    Args:
        data: List of dictionaries or objects with to_dict method
        columns: List of column names or dict {header: key}
        filename: Optional filename for download
    
    Returns:
        Flask Response with Excel data
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    
    # Handle column format
    if isinstance(columns, dict):
        headers = list(columns.keys())
        keys = list(columns.values())
    else:
        headers = columns
        keys = columns
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write data
    for row_num, item in enumerate(data, 2):
        if hasattr(item, 'to_dict'):
            item = item.to_dict()
        
        for col_num, key in enumerate(keys, 1):
            value = item.get(key, '')
            # Handle datetime objects
            if isinstance(value, datetime):
                value = value.isoformat()
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={filename or "export.xlsx"}'
        }
    )
    
    return response
